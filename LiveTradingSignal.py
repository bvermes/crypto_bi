#!/usr/bin/env python
# coding: utf-8

# # Init

# In[4]:


from data.datasource.tv_data_feed_api import get_data
from tvDatafeed import TvDatafeed, Interval
import requests

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

import pandas as pd
import matplotlib.pyplot as plt

import business.utils.trading_signals as ts
import business.utils.trading_indicators as ti

LOG_FILE = r"F:\Desktop\git\crypto_bi\log.xlsx"


# historical_data = get_data(symbol="BTCUSDT", exchange="BINANCE", interval=Interval.in_1_hour, n_bars=100000)
# historical_data


# In[5]:


def send_notification(message):
    # Replace these with your actual Pushover API token and user key
    app_token = 'acgrip2rbsfh6cdnghafcc922btrht'  # e.g., "a1b2c3d4e5f6g7h8i9j0"
    user_key = 'ucj8xoq8a275k5g7cn6nh91sdfk53m'    # e.g., "u1v2w3x4y5z6"

    # Prepare the payload with a loud sound ('siren' is one example)
    payload = {
        'token': app_token,
        'user': user_key,
        'message': message,
        'sound': 'siren',  # Change this to any sound of your choice
        'priority': 1      # priority 1 or 2 can trigger an alert mode on some devices
    }
    
    # Send the POST request to Pushover
    response = requests.post("https://api.pushover.net/1/messages.json", data=payload)
    
    if response.status_code == 200:
        print("Notification sent successfully!")
    else:
        print("Failed to send notification. Response:", response.text)


# In[6]:


def log_to_excel(timestamp, symbol, status, message, log_file=LOG_FILE):
    # If the log file doesn't exist, create it with headers
    if not os.path.exists(log_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        # Write header row
        ws.append(["Timestamp", "Symbol", "Status", "Message"])
        wb.save(log_file)
    
    # Load the existing workbook and append a new row
    wb = load_workbook(log_file)
    ws = wb.active
    ws.append([timestamp, symbol, status, message])
    wb.save(log_file)


# # Trading signals

# In[ ]:





# In[ ]:





# # Sandbox

# In[ ]:


df = get_data(
            symbol='ETHUSDT',
            exchange="BINANCE",
            interval=Interval.in_1_hour,
            n_bars=100000
        )


# In[24]:


df.shape


# In[23]:


df.head()


# In[9]:


df['RSI'], df['RSI_MA'] = ti.calculate_rsi_with_ma(df['close'], rsi_period=14, ma_type="SMA", ma_length=14)
df['MACD'], df['Signal'], df['MACD_Hist'] = ti.calculate_macd(df['close'], fast_period=12, slow_period=26, signal_period=9)
df['BB_Mid'], df['BB_Upper'], df['BB_Lower'] = ti.calculate_bollinger_bands(df['close'], window=20, num_std=2)
df['Stoch_K'], df['Stoch_D'] = ti.calculate_stochastic(df, k_period=14, d_period=3)


# Create subplots: one for the price and one for the RSI and its MA
fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8), sharex=True)

# Top subplot: Plot the closing prices
ax1.plot(df.index, df['close'], label='Closing Price', color='blue')
ax1.set_ylabel('Price')
ax1.set_title('Stock Price')
ax1.legend()
ax1.grid(True)

# Bottom subplot: Plot the RSI and its moving average
ax2.plot(df.index, df['RSI'], label='RSI', color='red')
ax2.plot(df.index, df['RSI_MA'], label='RSI MA', color='orange')
ax2.set_ylabel('RSI')
ax2.set_xlabel('Day')
ax2.set_title('RSI (14-period) with Moving Average')
ax2.legend()
ax2.grid(True)

# Adjust layout to prevent overlapping and display the plot
plt.tight_layout()
plt.show()


# In[10]:


fig, (ax1, ax2, ax3, ax4) = plt.subplots(4, 1, figsize=(12, 12), sharex=True)

# --- Panel 1: Price with Bollinger Bands ---
ax1.plot(df.index, df['close'], label='Closing Price', color='blue')
ax1.plot(df.index, df['BB_Mid'], label='BB Mid', color='black', linestyle='--')
ax1.plot(df.index, df['BB_Upper'], label='BB Upper', color='green', linestyle='--')
ax1.plot(df.index, df['BB_Lower'], label='BB Lower', color='red', linestyle='--')
ax1.set_ylabel('Price')
ax1.set_title('Stock Price with Bollinger Bands')
ax1.legend()
ax1.grid(True)

# --- Panel 2: RSI with its Moving Average ---
ax2.plot(df.index, df['RSI'], label='RSI', color='red')
ax2.plot(df.index, df['RSI_MA'], label='RSI MA', color='orange')
ax2.set_ylabel('RSI')
ax2.set_title('RSI (14) with Moving Average')
ax2.legend()
ax2.grid(True)

# --- Panel 3: MACD ---
ax3.plot(df.index, df['MACD'], label='MACD', color='blue')
ax3.plot(df.index, df['Signal'], label='Signal', color='orange')
# Plot histogram as bars
ax3.bar(df.index, df['MACD_Hist'], label='Histogram', color='grey', alpha=0.3)
ax3.set_ylabel('MACD')
ax3.set_title('MACD (12,26,9)')
ax3.legend()
ax3.grid(True)

# --- Panel 4: Stochastic Oscillator ---
ax4.plot(df.index, df['Stoch_K'], label='%K', color='purple')
ax4.plot(df.index, df['Stoch_D'], label='%D', color='brown')
ax4.set_ylabel('Stochastic')
ax4.set_title('Stochastic Oscillator (14,3)')
ax4.set_xlabel('Date')
ax4.legend()
ax4.grid(True)

plt.tight_layout()
plt.show()


# In[14]:


df.iloc[45:50].head()


# # Live Signal Multiple Stocks

# In[4]:


# List of symbols to monitor
symbols = ["BTCUSDT", "ETHUSDT"]

check_signals = [ts.very_low_price,
                    #very_high_price
                    ]


# In[5]:


def run_checks_and_notify(price_threshold):
    for symbol in symbols:
        # Retrieve data for the symbol (adjust parameters as needed)
        df = get_data(
            symbol=symbol,
            exchange="BINANCE",
            interval=Interval.in_1_hour,
            n_bars=100
        )
        
        # Run each check function on the data
        for check_func in check_signals:
            status, message = check_func(symbol, df, price_threshold)
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_to_excel(current_time, symbol, status, message)
            if status:
                send_notification(message)


# In[6]:


run_checks_and_notify(price_threshold=2560)

